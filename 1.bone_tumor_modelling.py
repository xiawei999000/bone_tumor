import os
import pandas as pd
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
from torchvision import models, transforms
from sklearn.model_selection import StratifiedShuffleSplit
from torchvision.models import resnet50,  inception_v3
from sklearn.metrics import roc_auc_score
from openpyxl import Workbook, load_workbook
from PIL import Image
import random

class FocalLoss(nn.Module):
    def __init__(self, alpha=0.25, gamma=2, reduction='mean'):
        super(FocalLoss, self).__init__()
        self.alpha = alpha
        self.gamma = gamma
        self.reduction = reduction

    def forward(self, inputs, targets):
        pt = inputs * targets + (1 - inputs) * (1 - targets)
        focal_loss = -self.alpha * (1 - pt) ** self.gamma * torch.log(pt + 1e-8)
        if self.reduction == 'mean':
            return focal_loss.mean()
        elif self.reduction == 'sum':
            return focal_loss.sum()
        else:
            return focal_loss
def write_to_global_excel(file_name, seed, learning_rate, batch_size, epoch,
                          val_auc, internal_test_auc, external_test_auc):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(['Seed', 'Learning Rate', 'Batch Size', 'epoch',
                   'Validation AUC', 'Internal Test AUC', 'External Test AUC'])
    else:
        wb = load_workbook(file_name)
        ws = wb.active

    ws.append([seed, learning_rate, batch_size, epoch,
               val_auc, internal_test_auc, external_test_auc])

    wb.save(file_name)

class BoneTumorDataset(Dataset):
    def __init__(self, image_dir, df, transform=None):
        self.df = df
        self.transform = transform

        self.images = []
        for img_name in df['image_id']:
            img_path = os.path.join(image_dir, img_name)
            image = Image.open(img_path).convert('RGB')
            if self.transform:
                image = self.transform(image)
            self.images.append(image)

    def __len__(self):
        return len(self.df)

    def __getitem__(self, idx):
        image = self.images[idx]
        label = self.df.iloc[idx]['label']
        return image, label

def calculate_auc(loader, model, device):
    model.eval()
    all_labels = []
    all_probs = []

    with torch.no_grad():
        for images, labels in loader:
            images, labels = images.to(device), labels.to(device)
            outputs = model(images)
            outputs = outputs.squeeze()
            probs = outputs.cpu().numpy()
            probs = np.round(probs, 4)
            all_probs.extend(probs)
            all_labels.extend(labels.cpu().numpy())

    auc = roc_auc_score(all_labels, all_probs)
    return auc

class Classifier(nn.Module):
        def __init__(self, num_class):
            super().__init__()
            self.drop_out = nn.Dropout(p=0.5)
            self.linear = nn.Linear(2048, num_class)
            self.Sigmoid = nn.Sigmoid()

        def forward(self, x):
            x = x.view(x.size(0), -1)
            x = self.drop_out(x)
            x = self.linear(x)
            # x = torch.softmax(x, dim=-1)
            x = self.Sigmoid(x)
            return x

class Backbone_ResNet50(nn.Module):
    def __init__(self):
        super().__init__()
        base_model = resnet50(pretrained=False)
        encoder_layers = list(base_model.children())
        self.backbone = nn.Sequential(*encoder_layers[:9])

    def forward(self, x):
        return self.backbone(x)
class Backbone_InceptionV3(nn.Module):
    def __init__(self, path):
        super().__init__()
        base_model = inception_v3(pretrained=False, aux_logits=False)
        encoder_layers = list(base_model.children())
        self.backbone = nn.Sequential(*encoder_layers[:-1])

        state_dict = torch.load(path)
        new_state_dict = {}
        for k, v in state_dict.items():
            new_state_dict[k[9:]] = v

        print(self.backbone.load_state_dict(new_state_dict))  # <All keys matched successfully>

    def forward(self, x):
        return self.backbone(x)

def main(model_type, seed, batch_size, learning_rate, num_epochs, global_results_file, pretrained_model_path, output_dir):

    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    random.seed(seed)

    rj_image_dir = './jpgs_patches_RJ-adjust'
    btxrd_image_dir = './jpgs_patches_BTXRD'
    rj_excel_path = './dataset_RJ-adjust.xlsx'
    btxrd_excel_path = './dataset_tumor_BTXRD-adjust.xlsx'

    rj_df = pd.read_excel(rj_excel_path)
    btxrd_df = pd.read_excel(btxrd_excel_path)

    rj_df['image_id'] = rj_df['image_id'].apply(lambda x: f"{x}.jpg")
    btxrd_df['image_id'] = btxrd_df['image_id'].apply(lambda x: f"{x}.jpeg")

    splitter = StratifiedShuffleSplit(n_splits=1, test_size=0.4, random_state=seed)
    train_idx, temp_idx = next(splitter.split(rj_df, rj_df['label']))

    train_df = rj_df.iloc[train_idx]
    temp_df = rj_df.iloc[temp_idx]

    splitter = StratifiedShuffleSplit(n_splits=1, test_size=0.5, random_state=seed)
    val_idx, test_idx = next(splitter.split(temp_df, temp_df['label']))

    val_df = temp_df.iloc[val_idx]
    internal_test_df = temp_df.iloc[test_idx]

    external_test_df = btxrd_df

    # for training
    transform = transforms.Compose([
        transforms.RandomHorizontalFlip(),
        transforms.RandomVerticalFlip(),
        transforms.RandomAffine(degrees=10, translate=(0.2, 0.2), scale=(0.8, 1.2)),
        transforms.CenterCrop(224),
        transforms.ToTensor()
    ])

    if model_type == 'InceptionV3_ImageNet':
        transform = transforms.Compose([
            transforms.RandomHorizontalFlip(),
            transforms.RandomVerticalFlip(),
            transforms.RandomAffine(degrees=10, translate=(0.2, 0.2), scale=(0.8, 1.2)),
            transforms.Resize(299),
            transforms.CenterCrop(299),
            transforms.ToTensor()
        ])

    # for validation and test
    transform_val = transforms.Compose([
        transforms.ToTensor()
    ])

    if model_type == 'InceptionV3_ImageNet':
        transform_val = transforms.Compose([
            transforms.Resize(299),
            transforms.CenterCrop(299),
            transforms.ToTensor()
        ])

    train_dataset = BoneTumorDataset(rj_image_dir, train_df, transform=transform)
    val_dataset = BoneTumorDataset(rj_image_dir, val_df, transform=transform_val)
    internal_test_dataset = BoneTumorDataset(rj_image_dir, internal_test_df, transform=transform_val)
    external_test_dataset = BoneTumorDataset(btxrd_image_dir, external_test_df, transform=transform_val)

    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=batch_size, shuffle=False)
    internal_test_loader = DataLoader(internal_test_dataset, batch_size=batch_size, shuffle=False)
    external_test_loader = DataLoader(external_test_dataset, batch_size=batch_size, shuffle=False)

    model = models.resnet50(pretrained=True)
    if model_type == "ResNet50_ImageNet":
        num_ftrs = model.fc.in_features
        model.fc = nn.Linear(num_ftrs, 1)
        model = nn.Sequential(model, nn.Sigmoid())

    if model_type == "InceptionV3_ImageNet":
        model = models.inception_v3(pretrained=True)
        model.aux_logits = False
        model.fc = nn.Linear(model.fc.in_features, 1)
        model = nn.Sequential(model, nn.Sigmoid())

    backbone = Backbone_ResNet50()
    if model_type == "ResNet50_RadImgNet":
        backbone.load_state_dict(torch.load(pretrained_model_path, weights_only=True))
        classifier = Classifier(num_class=1)
        model = nn.Sequential(backbone, classifier)

    if model_type == "InceptionV3_RadImgNet":
        backbone = Backbone_InceptionV3(pretrained_model_path)
        classifier = Classifier(num_class=1)
        model = nn.Sequential(backbone, classifier)

    criterion = FocalLoss()
    optimizer = optim.Adam(model.parameters(), lr=learning_rate)

    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(
        optimizer, mode='max', factor=0.618, patience=10, verbose=True
    )

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    model.to(device)

    best_val_auc = 0.5
    lambda_l2 = 1e-4

    num_epochs = num_epochs
    for epoch in range(num_epochs):
        model.train()
        torch.set_grad_enabled(True)
        running_loss = 0.0

        for images, labels in train_loader:
            images, labels = images.to(device), labels.to(device)

            optimizer.zero_grad()
            outputs = model(images)
            outputs = outputs.squeeze()

            loss = criterion(outputs, labels.float())
            l2_reg = torch.tensor(0.).to(images.device)
            for param in model.parameters():
                l2_reg += torch.norm(param, 2)
            loss += lambda_l2 * l2_reg

            running_loss += loss.item()

            loss.backward()
            optimizer.step()

        print(f"Epoch [{epoch+1}/{num_epochs}], Loss: {running_loss/len(train_loader):.4f}")

        model.eval()
        torch.set_grad_enabled(False)
        val_auc = calculate_auc(val_loader, model, device)
        internal_test_auc = calculate_auc(internal_test_loader, model, device)
        external_test_auc = calculate_auc(external_test_loader, model, device)

        val_auc = np.round(val_auc, 4)
        internal_test_auc = np.round(internal_test_auc,4)
        external_test_auc = np.round(external_test_auc, 4)

        print(f"Internal Validation AUC: {val_auc}, "
              f"Internal Test AUC: {internal_test_auc}, "
              f"External Test AUC: {external_test_auc}")

        scheduler.step(val_auc)

        current_lr = optimizer.param_groups[0]['lr']
        print(f"Current Learning Rate: {current_lr:.6f}")

        write_to_global_excel(global_results_file, seed, learning_rate, batch_size, epoch,
                              val_auc, internal_test_auc, external_test_auc)

        if val_auc > best_val_auc:
            best_val_auc = val_auc
            best_model_path = os.path.join(output_dir,
                                           f"best_model_seed_{seed}_lr_{learning_rate}_bs_{batch_size}_epc{epoch}_inAUC_{internal_test_auc:.3f}_exAUC_{external_test_auc:.3f}.pt")
            torch.save(model.state_dict(), best_model_path)
            print(f"Best model saved to {best_model_path}")


if __name__ == "__main__":

    seed = 666
    time = "250415-1657-1"

    num_epochs = 100
    batch_size = [128]
    learning_rate = 4e-3

    model_type_list = ["ResNet50_RadImgNet", "InceptionV3_RadImgNet",
                       "ResNet50_ImageNet", "InceptionV3_ImageNet"]

    for model_type in model_type_list:
        # pretrained model path (RadImgNet)
        pretrained_model_path = []
        if model_type == "ResNet50_RadImgNet":
            pretrained_model_path = "./RadImgNet/ResNet50.pt"
        if model_type == "InceptionV3_RadImgNet":
            pretrained_model_path = "./RadImgNet/InceptionV3.pt"

        output_dir = "./" + time + "_" + model_type
        os.makedirs(output_dir, exist_ok=True)
        global_results_file = os.path.join(output_dir, time + "_" + model_type + ".xlsx")


        print(f"Running with Seed={seed}, batch_size={batch_size}, Learning Rate={learning_rate}, num_epochs={num_epochs}")
        main(model_type, seed, batch_size, learning_rate, num_epochs, global_results_file, pretrained_model_path, output_dir)
